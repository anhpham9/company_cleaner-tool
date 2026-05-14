import pandas as pd
import unicodedata
import os
from openpyxl import load_workbook
from copy import copy


# ===== ✅ copy format nhanh (optimized) =====
def apply_format_fast(src_file, dst_file):

    wb_src = load_workbook(src_file)
    wb_dst = load_workbook(dst_file)

    ws_src = wb_src.active
    ws_dst = wb_dst.active

    # ✅ Copy column width
    for col_letter, col_dim in ws_src.column_dimensions.items():
        if col_dim.width:
            ws_dst.column_dimensions[col_letter].width = col_dim.width

    # ✅ Copy header style (row 1 only)
    for col in range(1, ws_src.max_column + 1):
        src_cell = ws_src.cell(row=1, column=col)
        dst_cell = ws_dst.cell(row=1, column=col)

        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)

    # ✅ Freeze pane
    ws_dst.freeze_panes = ws_src.freeze_panes

    wb_dst.save(dst_file)


# ===== ✅ normalize company name =====
def normalize(text):
    if pd.isna(text):
        return ""

    text = str(text)

    # ✅ remove /xxxx
    text = text.split("/")[0]

    # ✅ full-width → half-width
    text = unicodedata.normalize('NFKC', text)

    return text.strip().lower()


# ===== ✅ main process =====
def process_file(file_path, blacklist_path="blacklist.xlsx"):

    df = pd.read_excel(file_path)

    # ✅ giữ header trống (tránh Unnamed)
    df.columns = [
        "" if str(col).startswith("Unnamed") else col
        for col in df.columns
    ]

    # ✅ check cột company name
    if "会社名" not in df.columns:
        raise Exception("「会社名」列が見つかりません。")


    # ✅ clean company name để so sánh
    df["__clean"] = df["会社名"].apply(normalize)

    # ✅ update hiển thị (đã clean)
    df["会社名"] = df["会社名"].apply(
        lambda x: unicodedata.normalize('NFKC', str(x)).split("/")[0].strip()
        if pd.notna(x) else ""
    )

    # ===== blacklist =====
    blacklist_removed = pd.DataFrame()

    if os.path.exists(blacklist_path):
        bl = pd.read_excel(blacklist_path)
        bl["__clean"] = bl["会社名"].apply(normalize)

        # ✅ dòng bị blacklist
        blacklist_removed = df[df["__clean"].isin(bl["__clean"])].copy()
        blacklist_removed["削除理由"] = "ブラックリスト"

        # ✅ giữ lại dòng không thuộc blacklist
        df = df[~df["__clean"].isin(bl["__clean"])]


    # ===== duplicate =====
    duplicates = df[df.duplicated("__clean", keep="first")].copy()
    duplicates["削除理由"] = "重複"

    df_unique = df.drop_duplicates("__clean", keep="first")

    
    # ✅ removed count đúng
    removed = len(duplicates) + len(blacklist_removed)
    remain = len(df_unique)



    # ===== ✅ drop helper =====
    df_unique = df_unique.drop(columns=["__clean"])
    duplicates = duplicates.drop(columns=["__clean"])
    blacklist_removed = blacklist_removed.drop(columns=["__clean"], errors="ignore")


    base = os.path.splitext(file_path)[0]

    output_file = base + "_cleaned.xlsx"
    log_file = base + "_log.xlsx"
    df_unique.to_excel(output_file, index=False)
    
    # ✅ merge 2 loại log
    log_df = pd.concat([duplicates, blacklist_removed], ignore_index=True)
    
    if log_df.empty:
        log_df = pd.DataFrame(columns=list(df_unique.columns) + ["削除理由"])

    log_df.to_excel(log_file, index=False)


    # ✅ apply format FAST (QUAN TRỌNG)
    apply_format_fast(file_path, output_file)

    return {
        "removed": removed,
        "remain": remain,
        "output_file": output_file,
        "log_file": log_file
    }